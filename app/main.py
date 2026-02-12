import logging
import re
import time
import unicodedata
from decimal import Decimal
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Literal
from urllib.parse import parse_qsl, quote, urlencode

import anyio

from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import NoSuchTableError, OperationalError, ProgrammingError
from sqlalchemy.orm import Session
from sqlalchemy.sql import text
from starlette.requests import ClientDisconnect, Request

from app.db import get_session
from app.services.ingest import (
    IngestConflict,
    IngestError,
    IngestPersistenceError,
    ingest_excel,
)
from app.services.query import (
    build_series_query,
    get_availability,
    get_ingest_state,
    get_item_summary,
    get_latest_loaded_date,
    get_series_v2,
    get_suggestions,
    get_top_sales,
    resolve_project_groups,
)


def _normalize_csv_list(values: list[str] | None) -> list[str] | None:
    if not values:
        return None
    expanded: list[str] = []
    for value in values:
        if value is None:
            continue
        parts = [part.strip() for part in value.split(",")]
        expanded.extend(part for part in parts if part)
    return expanded or None


def _normalize_blank(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = value.strip()
    return stripped or None


def _parse_bool(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    normalized = value.strip().lower()
    if normalized in {"true", "1", "yes", "y", "on"}:
        return True
    if normalized in {"false", "0", "no", "n", "off", ""}:
        return False
    return default


def _parse_iso_date(value: date | str | None, label: str) -> date | None:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if not value.strip():
        return None
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail=f"{label} must be in YYYY-MM-DD format",
        ) from exc


def _format_ru_date(value: date | str | None) -> str:
    if value is None:
        return "—"
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    normalized = f"{value}".split("T")[0]
    try:
        return date.fromisoformat(normalized).strftime("%d.%m.%Y")
    except ValueError:
        return normalized


def _sanitize_filename(value: str, fallback: str = "export.xlsx") -> str:
    normalized = unicodedata.normalize("NFKC", value).strip()
    normalized = re.sub(r"[\\/:*?\"<>|]+", "_", normalized)
    normalized = re.sub(r"\s+", "_", normalized)
    normalized = normalized.strip("._")
    if not normalized:
        normalized = fallback
    if not normalized.lower().endswith(".xlsx"):
        normalized = f"{normalized}.xlsx"
    return normalized[:180]


def sanitize_ascii_filename(value: str, fallback: str = "export.xlsx") -> str:
    normalized = unicodedata.normalize("NFKC", value).strip()
    normalized = re.sub(r"[\\/:*?\"<>|]+", "_", normalized)
    normalized = re.sub(r"\s+", "_", normalized)
    normalized = "".join(char if ord(char) < 128 else "_" for char in normalized)
    normalized = normalized.strip("._")
    if not normalized:
        normalized = fallback
    if not normalized.lower().endswith(".xlsx"):
        normalized = f"{normalized}.xlsx"
    return normalized[:180]


_PRICE_FORMAT_INT = '# ##0 "₽"'
_PRICE_FORMAT_FLOAT = '# ##0,00 "₽"'
_MAX_COLUMN_WIDTH = 60
_COLUMN_WIDTH_PADDING = 2


def _format_price_for_width(value: int | float | Decimal) -> str:
    numeric = float(value)
    if numeric.is_integer():
        formatted = f"{int(round(numeric)):,}".replace(",", " ")
    else:
        formatted = f"{numeric:,.2f}".replace(",", " ").replace(".", ",")
    return f"{formatted} ₽"


def _cell_display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, WriteOnlyCell):
        cell_value = value.value
        if value.number_format in {_PRICE_FORMAT_INT, _PRICE_FORMAT_FLOAT} and isinstance(
            cell_value, (int, float, Decimal)
        ):
            return _format_price_for_width(cell_value)
        return _cell_display_value(cell_value)
    if isinstance(value, (date, datetime)):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, Decimal):
        return f"{value}"
    if isinstance(value, (str, int, float)):
        return str(value)
    logger.debug("Unexpected cell value type for width calc: %s", type(value))
    return str(value)


def _update_column_widths(
    widths: dict[int, int],
    row: list[object],
) -> None:
    for index, value in enumerate(row, start=1):
        try:
            length = len(_cell_display_value(value))
        except Exception:
            logger.exception("Failed to calculate cell width, falling back to str().")
            length = len("" if value is None else str(value))
        widths[index] = max(widths.get(index, 0), length)


def _append_row(
    worksheet,
    row: list[object],
    width_tracker: dict[str, dict[int, int]],
) -> None:
    worksheet.append(row)
    sheet_widths = width_tracker.setdefault(worksheet.title, {})
    _update_column_widths(sheet_widths, row)


def _apply_auto_width(
    workbook: Workbook,
    width_tracker: dict[str, dict[int, int]],
) -> None:
    try:
        for worksheet in workbook.worksheets:
            widths = width_tracker.get(worksheet.title, {})
            for index, max_len in widths.items():
                if max_len <= 0:
                    continue
                adjusted = min(
                    max(max_len + _COLUMN_WIDTH_PADDING, 8), _MAX_COLUMN_WIDTH
                )
                worksheet.column_dimensions[get_column_letter(index)].width = adjusted
    except Exception:
        logger.exception("Failed to apply auto-fit column widths.")


def _create_price_cell(worksheet, value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        cell = WriteOnlyCell(worksheet, value=value)
        if float(value).is_integer():
            cell.number_format = _PRICE_FORMAT_INT
        else:
            cell.number_format = _PRICE_FORMAT_FLOAT
        return cell
    return value


def _sanitize_sheet_name(name: str, existing: set[str]) -> str:
    clean = re.sub(r"[:\\/?*\[\]]", "_", name).strip() or "Sheet"
    clean = clean[:31]
    candidate = clean
    counter = 1
    while candidate in existing:
        suffix = f"_{counter}"
        candidate = f"{clean[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing.add(candidate)
    return candidate


def content_disposition(filename_unicode: str) -> str:
    ascii_fallback = sanitize_ascii_filename(filename_unicode)
    utf8_star = quote(filename_unicode, safe="")
    header_value = (
        f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{utf8_star}"
    )
    try:
        header_value.encode("ascii")
    except UnicodeEncodeError:
        logger.debug(
            "Content-Disposition header contains non-ascii characters",
            extra=safe_extra({"header": header_value}),
        )
    else:
        logger.debug(
            "Content-Disposition header generated",
            extra=safe_extra({"header": header_value}),
        )
    return header_value


def _normalize_query_string(query_string: bytes) -> bytes:
    if not query_string:
        return query_string
    params = parse_qsl(query_string.decode(), keep_blank_values=True)
    filtered = [
        (key, value.strip())
        for key, value in params
        if value is not None and value.strip() != ""
    ]
    if len(filtered) == len(params):
        return query_string
    return urlencode(filtered, doseq=True).encode()


def _is_missing_schema_error(exc: Exception) -> bool:
    if isinstance(exc, NoSuchTableError):
        return True
    orig = getattr(exc, "orig", None)
    if orig is not None:
        pgcode = getattr(orig, "pgcode", None)
        if pgcode == "42P01":
            return True
        if "undefined_table" in str(orig):
            return True
    message = str(exc).lower()
    return "no such table" in message or "relation" in message and "does not exist" in message


def _normalize_company(company: str | None, default: str | None = "alliance") -> str:
    raw_value = (company or "").strip()
    if not raw_value:
        if default is None:
            raise HTTPException(status_code=400, detail="Company is required")
        raw_value = default
    company_norm = raw_value.lower()
    company_alias = unicodedata.normalize("NFKC", company_norm).casefold()
    if company_alias in {"alliance", "альянс"}:
        company_norm = "alliance"
    allowed_companies = {"alliance", "vostok"}
    if company_norm not in allowed_companies:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown company: {company}",
        )
    return company_norm

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
RESERVED_LOG_KEYS = {
    "name",
    "msg",
    "args",
    "levelname",
    "levelno",
    "pathname",
    "filename",
    "module",
    "exc_info",
    "exc_text",
    "stack_info",
    "lineno",
    "funcName",
    "created",
    "msecs",
    "relativeCreated",
    "thread",
    "threadName",
    "processName",
    "process",
}


def safe_extra(extra: dict[str, object]) -> dict[str, object]:
    return {key: value for key, value in extra.items() if key not in RESERVED_LOG_KEYS}

app = FastAPI(title="Stock Delta Analyzer")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


@app.middleware("http")
async def normalize_query_params(request: Request, call_next):
    request.scope["query_string"] = _normalize_query_string(
        request.scope.get("query_string", b"")
    )
    return await call_next(request)


@app.middleware("http")
async def log_requests(request: Request, call_next):
    origin = request.headers.get("origin")
    logger.info("Request URL: %s; Origin: %s", request.url, origin)
    start_time = time.perf_counter()
    try:
        response = await call_next(request)
    except (anyio.EndOfStream, ClientDisconnect):
        duration_ms = (time.perf_counter() - start_time) * 1000
        logger.info(
            "Request completed",
            extra={
                "method": request.method,
                "path": request.url.path,
                "querystring": request.url.query,
                "status_code": 499,
                "duration_ms": round(duration_ms, 2),
            },
        )
        logger.info("Client disconnected during request body read: %s", request.url)
        return JSONResponse(
            status_code=499,
            content={"detail": "Client disconnected"},
        )
    except Exception:
        duration_ms = (time.perf_counter() - start_time) * 1000
        logger.info(
            "Request completed",
            extra={
                "method": request.method,
                "path": request.url.path,
                "querystring": request.url.query,
                "status_code": 500,
                "duration_ms": round(duration_ms, 2),
            },
        )
        logger.exception("Unhandled exception for %s", request.url)
        raise
    status_code = response.status_code
    duration_ms = (time.perf_counter() - start_time) * 1000
    logger.info(
        "Request completed",
        extra={
            "method": request.method,
            "path": request.url.path,
            "querystring": request.url.query,
            "status_code": status_code,
            "duration_ms": round(duration_ms, 2),
        },
    )
    if status_code >= 500:
        logger.error("Response status %s for %s", status_code, request.url)
    elif status_code >= 400:
        logger.warning("Response status %s for %s", status_code, request.url)
    else:
        logger.info("Response status %s for %s", status_code, request.url)
    return response


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/upload")
def upload_file(
    upload_date: date = Form(...),
    mode: Literal["reject", "bootstrap"] = Form("reject"),
    dry_run: bool = Query(False),
    company: str | None = Form(None),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    current_db = session.execute(text("select current_database()")).scalar()
    logger.debug("Current database: %s", current_db)
    try:
        file_bytes = file.file.read()
    except (ClientDisconnect, anyio.EndOfStream):
        return JSONResponse(
            status_code=499,
            content={"detail": "Client disconnected"},
        )
    logger.info("Upload received: %s bytes for %s", len(file_bytes), file.filename)
    company_norm = _normalize_company(company)
    try:
        payload = ingest_excel(
            session=session,
            upload_date=upload_date,
            file_bytes=file_bytes,
            company=company_norm,
            file_name=file.filename,
            mode=mode,
            dry_run=dry_run,
        )
    except (NoSuchTableError, OperationalError, ProgrammingError) as exc:
        if _is_missing_schema_error(exc):
            logger.exception("Database schema not migrated.")
            raise HTTPException(
                status_code=503,
                detail="DB schema not migrated",
            ) from exc
        raise
    except IngestConflict as exc:
        logger.warning("Upload conflict: %s", exc)
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except IngestPersistenceError as exc:
        logger.error("Upload persistence failure: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except IngestError as exc:
        logger.warning("Upload failed: %s", exc)
        if exc.report is not None:
            detail = {"message": str(exc), "validation_report": exc.report}
        else:
            detail = str(exc)
        raise HTTPException(status_code=400, detail=detail) from exc
    return payload


@app.get("/series")
def series(
    request: Request,
    item_id: int = Query(...),
    warehouses: list[str] | None = Query(default=None, alias="warehouses"),
    company: str | None = Query(default="alliance"),
    date_from: date = Query(...),
    date_to: date = Query(...),
    project_preset: str | None = Query(default=None),
    name_preset: str | None = Query(default=None),
    spring_subpreset: str | None = Query(default=None),
    session: Session = Depends(get_session),
):
    warehouses = _normalize_csv_list(warehouses)
    company_norm = _normalize_company(_normalize_blank(company))
    project_groups = resolve_project_groups(_normalize_blank(project_preset))
    name_preset = _normalize_blank(name_preset)
    spring_subpreset = _normalize_blank(spring_subpreset)
    if name_preset not in {None, "spring", "spring_extra"}:
        raise HTTPException(status_code=400, detail="name_preset must be one of: spring, spring_extra")
    if name_preset == "spring" and spring_subpreset is None:
        raise HTTPException(
            status_code=400,
            detail="spring_subpreset is required for name_preset='spring'",
        )
    if name_preset == "spring" and spring_subpreset not in {"leaf", "bushing", "u_bolt", "spring"}:
        raise HTTPException(
            status_code=400,
            detail="spring_subpreset must be one of: leaf, bushing, u_bolt, spring",
        )
    if name_preset != "spring":
        spring_subpreset = None
    logger.info(
        "Series preset params: name_preset=%s spring_subpreset=%s",
        name_preset,
        spring_subpreset,
    )
    logger.info("Series request query_string=%s", request.url.query)
    logger.info(
        "Series params normalized",
        extra={
            "company_norm": company_norm,
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "warehouses": warehouses,
            "item_id": item_id,
            "project_preset": _normalize_blank(project_preset),
            "name_preset": name_preset,
            "spring_subpreset": spring_subpreset,
            "spring_filter_applied": name_preset in {"spring", "spring_extra"},
        },
    )
    payload = get_series_v2(
        session=session,
        item_id=item_id,
        company=company_norm,
        warehouses=warehouses,
        date_from=date_from,
        date_to=date_to,
        project_groups=project_groups,
        name_preset=name_preset,
        spring_subpreset=spring_subpreset,
    )
    logger.info(
        "Series response",
        extra={
            "rows_count": len(payload.get("series", [])),
            "item_id": item_id,
        },
    )
    return payload


@app.get("/export/series")
def export_series(
    item_id: int = Query(...),
    warehouses: list[str] | None = Query(default=None, alias="warehouses"),
    company: str | None = Query(default="alliance"),
    date_from: date = Query(...),
    date_to: date = Query(...),
    project_preset: str | None = Query(default=None),
    name_preset: str | None = Query(default=None),
    spring_subpreset: str | None = Query(default=None),
    group_by_warehouse: str | None = Query(default=None),
    session: Session = Depends(get_session),
):
    start_time = time.perf_counter()
    warehouses = _normalize_csv_list(warehouses)
    company_norm = _normalize_company(_normalize_blank(company))
    project_groups = resolve_project_groups(_normalize_blank(project_preset))
    name_preset = _normalize_blank(name_preset)
    spring_subpreset = _normalize_blank(spring_subpreset)
    if name_preset not in {None, "spring", "spring_extra"}:
        raise HTTPException(status_code=400, detail="name_preset must be one of: spring, spring_extra")
    if name_preset == "spring" and spring_subpreset is None:
        raise HTTPException(status_code=400, detail="spring_subpreset is required for name_preset='spring'")
    if name_preset == "spring" and spring_subpreset not in {"leaf", "bushing", "u_bolt", "spring"}:
        raise HTTPException(status_code=400, detail="spring_subpreset must be one of: leaf, bushing, u_bolt, spring")
    if name_preset != "spring":
        spring_subpreset = None
    group_by_warehouse_flag = _parse_bool(group_by_warehouse, default=True)
    item_summary = get_item_summary(session=session, item_id=item_id, company=company_norm)
    if not item_summary:
        raise HTTPException(status_code=404, detail="Товар не найден.")

    stmt, resolved_warehouses, availability_range = build_series_query(
        session=session,
        item_id=item_id,
        company=company_norm,
        warehouses=warehouses,
        date_from=date_from,
        date_to=date_to,
        project_groups=project_groups,
        name_preset=name_preset,
        spring_subpreset=spring_subpreset,
    )
    if stmt is None:
        raise HTTPException(status_code=404, detail="Данные для серии не найдены.")

    workbook = Workbook(write_only=True)
    summary_sheet = workbook.create_sheet("Сводка")
    data_sheet = workbook.create_sheet("Данные")
    column_widths: dict[str, dict[int, int]] = {}
    _append_row(
        data_sheet,
        [
            "Дата",
            "Склад",
            "Остаток",
            "Продано",
            "Пополнено",
            "Цена",
        ],
        column_widths,
    )

    existing_sheet_names = {"Сводка", "Данные"}
    warehouse_sheets: dict[str, object] = {}
    include_aggregate = not group_by_warehouse_flag
    aggregated_by_date: dict[date, dict[str, float]] = {}
    aggregated_price_counts: dict[date, int] = {}
    sold_total = 0
    replenished_total = 0
    latest_by_warehouse: dict[str, dict[str, object]] = {}
    last_date: date | None = None
    rows_count = 0
    multiple_warehouses = len(resolved_warehouses) > 1

    for row in session.execute(stmt).mappings():
        data_date = row["data_date"]
        warehouse = row["warehouse"] or "—"
        stock_qty = int(row["stock_qty"]) if row["stock_qty"] is not None else 0
        sold_qty = int(row["sold_qty"] or 0)
        replenished_qty = int(row["replenished_qty"] or 0)
        price_value = float(row["price"]) if row["price"] is not None else None
        date_label = _format_ru_date(data_date)
        price_cell = _create_price_cell(data_sheet, price_value)
        _append_row(
            data_sheet,
            [date_label, warehouse, stock_qty, sold_qty, replenished_qty, price_cell],
            column_widths,
        )
        if multiple_warehouses:
            sheet = warehouse_sheets.get(warehouse)
            if sheet is None:
                sheet_name = _sanitize_sheet_name(warehouse, existing_sheet_names)
                sheet = workbook.create_sheet(sheet_name)
                _append_row(
                    sheet,
                    [
                        "Дата",
                        "Склад",
                        "Остаток",
                        "Продано",
                        "Пополнено",
                        "Цена",
                    ],
                    column_widths,
                )
                warehouse_sheets[warehouse] = sheet
            price_cell = _create_price_cell(sheet, price_value)
            _append_row(
                sheet,
                [date_label, warehouse, stock_qty, sold_qty, replenished_qty, price_cell],
                column_widths,
            )

        sold_total += sold_qty
        replenished_total += replenished_qty
        latest_by_warehouse[warehouse] = {
            "date": data_date,
            "stock_qty": stock_qty,
            "price": price_value,
        }
        if last_date is None or data_date > last_date:
            last_date = data_date

        if include_aggregate:
            summary = aggregated_by_date.setdefault(
                data_date,
                {"stock_qty": 0.0, "sold_qty": 0.0, "replenished_qty": 0.0, "price_sum": 0.0},
            )
            summary["stock_qty"] += stock_qty
            summary["sold_qty"] += sold_qty
            summary["replenished_qty"] += replenished_qty
            if price_value is not None:
                summary["price_sum"] += price_value
                aggregated_price_counts[data_date] = (
                    aggregated_price_counts.get(data_date, 0) + 1
                )

        rows_count += 1

    latest_stock_total = sum(
        entry.get("stock_qty", 0) for entry in latest_by_warehouse.values()
    )
    latest_prices = [
        entry.get("price")
        for entry in latest_by_warehouse.values()
        if entry.get("price") is not None
    ]
    if latest_prices:
        min_price = min(latest_prices)
        max_price = max(latest_prices)
        last_price_value: str | float = (
            float(min_price) if min_price == max_price else f"{min_price:.2f} – {max_price:.2f}"
        )
    else:
        last_price_value = "—"

    warehouses_label = (
        ", ".join(resolved_warehouses) if resolved_warehouses else "ВСЕ"
    )
    _append_row(
        summary_sheet,
        ["Артикул", item_summary.get("canonical_sku") or "—"],
        column_widths,
    )
    _append_row(
        summary_sheet,
        ["Наименование", item_summary.get("name") or "—"],
        column_widths,
    )
    _append_row(
        summary_sheet,
        ["Производитель", item_summary.get("manufacturer") or "—"],
        column_widths,
    )
    _append_row(summary_sheet, ["Компания", company_norm], column_widths)
    _append_row(
        summary_sheet,
        [
            "Период",
            f"{_format_ru_date(date_from)} - {_format_ru_date(date_to)}",
        ],
        column_widths,
    )
    _append_row(summary_sheet, ["Склад(ы)", warehouses_label], column_widths)
    _append_row(summary_sheet, [], column_widths)
    _append_row(summary_sheet, ["Продано", sold_total], column_widths)
    _append_row(summary_sheet, ["Пополнено", replenished_total], column_widths)
    _append_row(summary_sheet, ["Последняя цена", last_price_value], column_widths)
    _append_row(
        summary_sheet,
        ["Остаток на конец периода", latest_stock_total],
        column_widths,
    )
    _append_row(
        summary_sheet,
        [
            "Последняя загруженная дата",
            _format_ru_date(get_latest_loaded_date(session=session, company=company_norm)),
        ],
        column_widths,
    )
    _append_row(
        summary_sheet,
        [
            "Доступный период",
            f"{_format_ru_date(availability_range.get('min'))} - {_format_ru_date(availability_range.get('max'))}",
        ],
        column_widths,
    )

    if include_aggregate:
        aggregate_sheet = workbook.create_sheet("Итого")
        _append_row(
            aggregate_sheet,
            ["Примечание", "Цена = средняя по складам за дату"],
            column_widths,
        )
        _append_row(aggregate_sheet, [], column_widths)
        _append_row(
            aggregate_sheet,
            ["Дата", "Остаток", "Продано", "Пополнено", "Цена"],
            column_widths,
        )
        for agg_date in sorted(aggregated_by_date.keys()):
            summary = aggregated_by_date[agg_date]
            price_count = aggregated_price_counts.get(agg_date, 0)
            avg_price = (
                summary["price_sum"] / price_count if price_count > 0 else None
            )
            avg_price_cell = _create_price_cell(aggregate_sheet, avg_price)
            _append_row(
                aggregate_sheet,
                [
                    _format_ru_date(agg_date),
                    summary["stock_qty"],
                    summary["sold_qty"],
                    summary["replenished_qty"],
                    avg_price_cell,
                ],
                column_widths,
            )

    filename = _sanitize_filename(
        f"Товар_{item_summary.get('canonical_sku')}_{date_from.isoformat()}-{date_to.isoformat()}",
        fallback="series.xlsx",
    )
    output = BytesIO()
    _apply_auto_width(workbook, column_widths)
    workbook.save(output)
    output.seek(0)
    duration_ms = (time.perf_counter() - start_time) * 1000
    logger.info(
        "Series export generated",
        extra=safe_extra(
            {
                "item_id": item_id,
                "company": company_norm,
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
                "warehouses": warehouses,
                "rows_count": rows_count,
                "duration_ms": round(duration_ms, 2),
                "group_by_warehouse": group_by_warehouse_flag,
            }
        ),
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition(filename)},
    )


@app.get("/filters/suggestions")
def filter_suggestions(
    field: str = Query(..., description="sku, warehouse, manufacturer, name"),
    q: str = Query(""),
    company: str | None = Query(default="alliance"),
    session: Session = Depends(get_session),
):
    company_norm = _normalize_company(_normalize_blank(company))
    return {
        "items": get_suggestions(
            session=session,
            field=field,
            query=q,
            company=company_norm,
        )
    }


@app.get("/top")
def top_sales(
    request: Request,
    limit: int = Query(100, ge=1, le=2000),
    page: int = Query(1, ge=1),
    offset: int | None = Query(default=None, ge=0),
    company: str | None = Query(default="alliance"),
    sku: str | None = Query(default=None),
    manufacturer: str | None = Query(default=None),
    name: str | None = Query(default=None),
    project_label: str | None = Query(default=None, alias="project"),
    project_preset: str | None = Query(default=None),
    name_preset: str | None = Query(default=None),
    spring_subpreset: str | None = Query(default=None),
    q: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    session: Session = Depends(get_session),
):
    warehouses_raw = request.query_params.get("warehouses")
    warehouses = (
        [value.strip() for value in warehouses_raw.split(",") if value.strip()]
        if warehouses_raw
        else None
    )
    group_by_raw = request.query_params.get("group_by_warehouse")
    group_by_warehouse = _parse_bool(group_by_raw, default=False)
    sku = _normalize_blank(sku)
    manufacturer = _normalize_blank(manufacturer)
    project_label = _normalize_blank(project_label)
    name = _normalize_blank(name)
    project_preset = _normalize_blank(project_preset)
    name_preset = _normalize_blank(name_preset)
    spring_subpreset = _normalize_blank(spring_subpreset)
    q = _normalize_blank(q)
    if name_preset not in {None, "spring", "spring_extra"}:
        raise HTTPException(status_code=400, detail="name_preset must be one of: spring, spring_extra")
    if name_preset == "spring" and spring_subpreset is None:
        raise HTTPException(
            status_code=400,
            detail="spring_subpreset is required for name_preset='spring'",
        )
    if name_preset == "spring" and spring_subpreset not in {"leaf", "bushing", "u_bolt", "spring"}:
        raise HTTPException(
            status_code=400,
            detail="spring_subpreset must be one of: leaf, bushing, u_bolt, spring",
        )
    if name_preset != "spring":
        spring_subpreset = None
    if project_label:
        project_preset = None
    project_groups = resolve_project_groups(project_preset)
    company_norm = _normalize_company(_normalize_blank(company))
    date_from = _parse_iso_date(date_from, "date_from")
    date_to = _parse_iso_date(date_to, "date_to")
    if date_from is None and date_to is None:
        date_to = date.today()
        date_from = date_to - timedelta(days=30)
    if date_from is None and date_to is not None:
        date_from = date_to
    if date_to is None and date_from is not None:
        date_to = date_from
    resolved_offset = offset if offset is not None else (page - 1) * limit
    logger.info(
        "Top params q=%r page=%s limit=%s company=%s warehouses=%s date_from=%s date_to=%s project=%s project_preset=%s name_preset=%s spring_subpreset=%s group_by_warehouse=%s",
        q,
        page,
        limit,
        company_norm,
        warehouses,
        date_from,
        date_to,
        project_label,
        project_preset,
        name_preset,
        spring_subpreset,
        group_by_warehouse,
    )
    logger.info(
        "Top request filters: limit=%s page=%s company=%s warehouses=%s date_from=%s date_to=%s",
        limit,
        page,
        company_norm,
        warehouses,
        date_from,
        date_to,
    )
    logger.info(
        "Top preset params: name_preset=%s spring_subpreset=%s",
        name_preset,
        spring_subpreset,
    )
    normalized_query_string = request.scope.get("query_string", b"").decode()
    logger.info("Top request query_string=%s", request.url.query)
    logger.info("Top request normalized_query_string=%s", normalized_query_string)
    logger.info("Top request q=%s", q)
    logger.info(
        "Top params",
        extra=safe_extra(
            {
                "company_norm": company_norm,
                "date_from": date_from.isoformat() if date_from else None,
                "date_to": date_to.isoformat() if date_to else None,
                "warehouses": warehouses,
                "limit": limit,
                "page": page,
                "offset": resolved_offset,
                "manufacturer": manufacturer,
                "sku": sku,
                "item_name": name,
                "project": project_label,
                "project_preset": project_preset,
                "name_preset": name_preset,
                "spring_subpreset": spring_subpreset,
                "spring_filter_applied": name_preset in {"spring", "spring_extra"},
                "group_by_warehouse": group_by_warehouse,
                "q": q,
            }
        ),
    )
    try:
        payload = get_top_sales(
            session=session,
            limit=limit,
            offset=resolved_offset,
            company=company_norm,
            warehouses=warehouses,
            sku=sku,
            manufacturer=manufacturer,
            name=name,
            project=project_label,
            project_groups=project_groups,
            group_by_warehouse=group_by_warehouse,
            date_from=date_from,
            date_to=date_to,
            name_preset=name_preset,
            spring_subpreset=spring_subpreset,
            q=q,
        )
    except Exception:
        logger.exception(
            "Top sales query failed",
            extra=safe_extra(
                {
                    "company_norm": company_norm,
                    "date_from": date_from.isoformat() if date_from else None,
                    "date_to": date_to.isoformat() if date_to else None,
                    "warehouses": warehouses,
                    "limit": limit,
                    "page": page,
                    "offset": resolved_offset,
                    "group_by_warehouse": group_by_warehouse,
                }
            ),
        )
        raise
    logger.info(
        "Top results",
        extra=safe_extra(
            {
                "q": q,
                "q_filter_applied": bool(q),
                "rows_count": len(payload.get("items", [])),
                "total_count": payload.get("total_count"),
            }
        ),
    )
    return {
        "items": payload.get("items", []),
        "total_count": payload.get("total_count", 0),
        "page": page,
        "limit": limit,
    }


@app.get("/export/top")
def export_top(
    request: Request,
    limit: int = Query(100, ge=1, le=2000),
    page: int = Query(1, ge=1),
    company: str | None = Query(default="alliance"),
    sku: str | None = Query(default=None),
    manufacturer: str | None = Query(default=None),
    name: str | None = Query(default=None),
    project_label: str | None = Query(default=None, alias="project"),
    project_preset: str | None = Query(default=None),
    name_preset: str | None = Query(default=None),
    spring_subpreset: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    export_all: str | None = Query(default=None),
    session: Session = Depends(get_session),
):
    start_time = time.perf_counter()
    warehouses_raw = request.query_params.get("warehouses")
    warehouses = (
        [value.strip() for value in warehouses_raw.split(",") if value.strip()]
        if warehouses_raw
        else None
    )
    group_by_raw = request.query_params.get("group_by_warehouse")
    group_by_warehouse = _parse_bool(group_by_raw, default=False)
    export_all_flag = _parse_bool(export_all, default=False)
    sku = _normalize_blank(sku)
    manufacturer = _normalize_blank(manufacturer)
    project_label = _normalize_blank(project_label)
    name = _normalize_blank(name)
    project_preset = _normalize_blank(project_preset)
    name_preset = _normalize_blank(name_preset)
    spring_subpreset = _normalize_blank(spring_subpreset)
    if name_preset not in {None, "spring", "spring_extra"}:
        raise HTTPException(status_code=400, detail="name_preset must be one of: spring, spring_extra")
    if name_preset == "spring" and spring_subpreset is None:
        raise HTTPException(status_code=400, detail="spring_subpreset is required for name_preset='spring'")
    if name_preset == "spring" and spring_subpreset not in {"leaf", "bushing", "u_bolt", "spring"}:
        raise HTTPException(status_code=400, detail="spring_subpreset must be one of: leaf, bushing, u_bolt, spring")
    if name_preset != "spring":
        spring_subpreset = None
    if project_label:
        project_preset = None
    project_groups = resolve_project_groups(project_preset)
    company_norm = _normalize_company(_normalize_blank(company))
    date_from = _parse_iso_date(date_from, "date_from")
    date_to = _parse_iso_date(date_to, "date_to")
    if date_from is None and date_to is None:
        date_to = date.today()
        date_from = date_to - timedelta(days=30)

    workbook = Workbook(write_only=True)
    top_sheet = workbook.create_sheet("Топ")
    column_widths: dict[str, dict[int, int]] = {}
    _append_row(
        top_sheet,
        [
            "Ранг",
            "Артикул",
            "Наименование",
            "Склад",
            "Продано",
            "Пополнено",
            "Последняя цена",
            "Группа/Проект",
        ],
        column_widths,
    )

    rows_written = 0
    total_count = 0
    export_cap = 20000
    if export_all_flag:
        total_payload = get_top_sales(
            session=session,
            limit=1,
            offset=0,
            company=company_norm,
            warehouses=warehouses,
            sku=sku,
            manufacturer=manufacturer,
            name=name,
            project=project_label,
            project_groups=project_groups,
            group_by_warehouse=group_by_warehouse,
            date_from=date_from,
            date_to=date_to,
            name_preset=name_preset,
            spring_subpreset=spring_subpreset,
        )
        total_count = total_payload.get("total_count", 0)
        if total_count > export_cap:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Слишком много строк для экспорта ({total_count}). "
                    "Уточните фильтры или выберите экспорт текущей страницы."
                ),
            )
        batch_size = min(1000, max(limit, 1))
        for offset in range(0, total_count, batch_size):
            payload = get_top_sales(
                session=session,
                limit=batch_size,
                offset=offset,
                company=company_norm,
                warehouses=warehouses,
                sku=sku,
                manufacturer=manufacturer,
                name=name,
                project=project_label,
                project_groups=project_groups,
                group_by_warehouse=group_by_warehouse,
                date_from=date_from,
                date_to=date_to,
                name_preset=name_preset,
                spring_subpreset=spring_subpreset,
            )
            items = payload.get("items", [])
            for item in items:
                warehouse_label = (
                    "ВСЕ"
                    if not group_by_warehouse
                    else (item.get("warehouse") or "—")
                )
                price_cell = _create_price_cell(top_sheet, item.get("last_price"))
                _append_row(
                    top_sheet,
                    [
                        item.get("rank") or "—",
                        item.get("canonical_sku") or "—",
                        item.get("name") or "—",
                        warehouse_label,
                        item.get("sold_total") or 0,
                        item.get("replenished_total") or 0,
                        price_cell,
                        item.get("group_name") or "—",
                    ],
                    column_widths,
                )
            rows_written += len(items)
    else:
        resolved_offset = (page - 1) * limit
        payload = get_top_sales(
            session=session,
            limit=limit,
            offset=resolved_offset,
            company=company_norm,
            warehouses=warehouses,
            sku=sku,
            manufacturer=manufacturer,
            name=name,
            project=project_label,
            project_groups=project_groups,
            group_by_warehouse=group_by_warehouse,
            date_from=date_from,
            date_to=date_to,
            name_preset=name_preset,
            spring_subpreset=spring_subpreset,
        )
        items = payload.get("items", [])
        total_count = payload.get("total_count", len(items))
        for item in items:
            warehouse_label = (
                "ВСЕ" if not group_by_warehouse else (item.get("warehouse") or "—")
            )
            price_cell = _create_price_cell(top_sheet, item.get("last_price"))
            _append_row(
                top_sheet,
                [
                    item.get("rank") or "—",
                    item.get("canonical_sku") or "—",
                    item.get("name") or "—",
                    warehouse_label,
                    item.get("sold_total") or 0,
                    item.get("replenished_total") or 0,
                    price_cell,
                    item.get("group_name") or "—",
                ],
                column_widths,
            )
        rows_written = len(items)

    params_sheet = workbook.create_sheet("Параметры")
    _append_row(params_sheet, ["Компания", company_norm], column_widths)
    _append_row(
        params_sheet,
        ["Период", f"{_format_ru_date(date_from)} - {_format_ru_date(date_to)}"],
        column_widths,
    )
    _append_row(
        params_sheet,
        ["Склады", ", ".join(warehouses) if warehouses else "ВСЕ"],
        column_widths,
    )
    _append_row(params_sheet, ["Проект", project_label or "—"], column_widths)
    _append_row(params_sheet, ["Preset", project_preset or "—"], column_widths)
    _append_row(
        params_sheet,
        ["Группировка по складам", "Да" if group_by_warehouse else "Нет"],
        column_widths,
    )
    _append_row(
        params_sheet,
        ["Экспорт", "Все строки" if export_all_flag else "Текущая страница"],
        column_widths,
    )
    _append_row(
        params_sheet,
        ["Экспортировано", datetime.now().strftime("%d.%m.%Y %H:%M")],
        column_widths,
    )

    filename = _sanitize_filename(
        f"Топ_{date_from.isoformat()}-{date_to.isoformat()}",
        fallback="top.xlsx",
    )
    output = BytesIO()
    _apply_auto_width(workbook, column_widths)
    workbook.save(output)
    output.seek(0)
    duration_ms = (time.perf_counter() - start_time) * 1000
    logger.info(
        "Top export generated",
        extra=safe_extra(
            {
                "company": company_norm,
                "date_from": date_from.isoformat() if date_from else None,
                "date_to": date_to.isoformat() if date_to else None,
                "warehouses": warehouses,
                "group_by_warehouse": group_by_warehouse,
                "export_all": export_all_flag,
                "rows_count": rows_written,
                "total_count": total_count,
                "duration_ms": round(duration_ms, 2),
            }
        ),
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition(filename)},
    )


@app.get("/availability")
def availability(
    company: str | None = Query(default="alliance"),
    session: Session = Depends(get_session),
):
    company_norm = _normalize_company(_normalize_blank(company))
    return get_availability(session=session, company=company_norm)


@app.get("/meta/latest_date")
def latest_loaded_date(
    company: str | None = Query(default="alliance"),
    session: Session = Depends(get_session),
):
    company_norm = _normalize_company(_normalize_blank(company))
    latest_date = get_latest_loaded_date(session=session, company=company_norm)
    logger.info(
        "Latest loaded date",
        extra=safe_extra({"company": company_norm, "latest_date": latest_date}),
    )
    return {"latest_date": latest_date}


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/health/db")
def health_db(session: Session = Depends(get_session)):
    tables_to_check = ("items", "fact_snapshot", "fact_delta_changes", "ingest_runs")
    try:
        session.execute(text("select 1"))
        tables = {
            table: session.execute(
                text("select to_regclass(:table_name)"),
                {"table_name": f"public.{table}"},
            ).scalar()
            is not None
            for table in tables_to_check
        }
    except (OperationalError, ProgrammingError) as exc:
        logger.exception("Database health check failed.")
        return {"status": "error", "detail": str(exc)}

    missing_tables = [table for table, exists in tables.items() if not exists]
    if missing_tables:
        return {"status": "error", "detail": f"Missing tables: {', '.join(missing_tables)}"}

    return {"status": "ok", "tables": tables}


@app.get("/debug/ingest_state")
def ingest_state(
    company: str = Query(...),
    limit: int = Query(30, ge=1, le=366),
    session: Session = Depends(get_session),
):
    company_norm = _normalize_company(_normalize_blank(company), default=None)
    return get_ingest_state(session=session, company=company_norm, limit=limit)
